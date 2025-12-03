'''飞书多维表格数据处理脚本（适配GitHub Actions）'''
import os
import json
from datetime import datetime
import lark_oapi as lark
from lark_oapi.api.bitable.v1 import *
from lark_oapi.api.drive.v1 import *
import requests
import pyexcel
import pandas as pd
import io

# ====================== 环境变量配置（从GitHub Actions环境读取） ======================
# 从环境变量读取核心配置（需在GitHub仓库Secrets/Workflow中配置）
APP_ID = os.getenv("FEISHU_APP_ID")
APP_SECRET = os.getenv("FEISHU_APP_SECRET")
DWBG_TOKEN = os.getenv("DWBG_TOKEN")
DWBG_TABLE_ID = os.getenv("DWBG_TABLE_ID")
ROW_ID = os.getenv("ROW_ID")

# 校验必要环境变量是否存在
def validate_environment():
    """校验运行所需的环境变量是否完整"""
    missing_vars = []
    if not APP_ID:
        missing_vars.append("FEISHU_APP_ID")
    if not APP_SECRET:
        missing_vars.append("FEISHU_APP_SECRET")
    if not DWBG_TOKEN:
        missing_vars.append("DWBG_TOKEN")
    if not DWBG_TABLE_ID:
        missing_vars.append("DWBG_TABLE_ID")
    if not ROW_ID:
        missing_vars.append("ROW_ID")
    
    if missing_vars:
        raise Exception(f"❌ 缺少必要环境变量：{', '.join(missing_vars)}\n请检查GitHub Actions的Secrets/Payload配置")

'''飞书多维表格核心函数'''
def 获取访问令牌(APP_ID, APP_SECRET):
    """获取飞书租户访问令牌"""
    url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal/"
    headers = {"Content-Type": "application/json"}
    data = {
        "app_id": APP_ID,
        "app_secret": APP_SECRET
    }
    try:
        response = requests.post(url, headers=headers, json=data, timeout=10)
        response.raise_for_status()  # 抛出HTTP错误
        response_data = response.json()

        if response_data.get("code") == 0:
            token = response_data.get("tenant_access_token")
            print(f"✅ 获取访问令牌成功（前20位）：{token[:20]}...")
            return token
        else:
            raise Exception(f"获取access_token失败: 错误码={response_data.get('code')}, 消息={response_data.get('msg')}")
    except requests.exceptions.RequestException as e:
        raise Exception(f"获取access_token网络请求失败: {str(e)}")

def 飞书上传素材(文件路径, DWBG_TOKEN, 应用ID, 应用密匙):
    """使用飞书官方SDK上传文件到多维表格（保留函数，兼容原有逻辑）"""
    if not os.path.exists(文件路径):
        print(f"错误：文件不存在 - {文件路径}")
        return None
    if not os.path.isfile(文件路径):
        print(f"错误：不是有效的文件 - {文件路径}")
        return None
    
    file_name = os.path.basename(文件路径)
    file_size = os.path.getsize(文件路径)
    print(f"准备上传文件: {file_name} (大小: {file_size} bytes)")
    
    if file_size > 20 * 1024 * 1024:
        print(f"错误：文件过大，超过20MB限制")
        return None
    
    client = lark.Client.builder() \
        .app_id(应用ID) \
        .app_secret(应用密匙) \
        .log_level(lark.LogLevel.DEBUG) \
        .build()
    
    try:
        with open(文件路径, "rb") as file:
            request: UploadAllMediaRequest = UploadAllMediaRequest.builder() \
                .request_body(UploadAllMediaRequestBody.builder()
                              .file_name(file_name)
                              .parent_type("bitable_file")
                              .parent_node(DWBG_TOKEN)
                              .size(str(file_size))
                              .file(file)
                              .build()) \
                .build()

            response: UploadAllMediaResponse = client.drive.v1.media.upload_all(request)

            if not response.success():
                error_msg = f"文件上传失败 - 代码: {response.code}, 消息: {response.msg}, 日志ID: {response.get_log_id()}"
                print(error_msg)
                if response.raw and response.raw.content:
                    try:
                        resp_content = json.loads(response.raw.content)
                        print("详细响应内容:")
                        print(json.dumps(resp_content, indent=4, ensure_ascii=False))
                    except:
                        print("响应内容解析失败:", response.raw.content)
                return None
            else:
                print("文件上传成功!")
                print("返回数据:", lark.JSON.marshal(response.data, indent=4))
                return response.data.file_token
    except Exception as e:
        print(f"上传过程发生错误: {str(e)}")
        return None

def 新增飞书表格(应用ID, 应用密匙, DWBG_TOKEN, DWBG_TABLE_ID, 上传数据结构):
    """新增飞书多维表格记录（保留函数，兼容原有逻辑）"""
    client = lark.Client.builder() \
        .app_id(应用ID) \
        .app_secret(应用密匙) \
        .log_level(lark.LogLevel.DEBUG) \
        .build()

    request: CreateAppTableRecordRequest = CreateAppTableRecordRequest.builder() \
        .app_token(DWBG_TOKEN) \
        .table_id(DWBG_TABLE_ID) \
        .request_body(AppTableRecord.builder()
                      .fields(上传数据结构)
                      .build()) \
        .build()

    response: CreateAppTableRecordResponse = client.bitable.v1.app_table_record.create(request)

    if not response.success():
        lark.logger.error(
            f"client.bitable.v1.app_table_record.create failed, code: {response.code}, msg: {response.msg}, log_id: {response.get_log_id()}, resp: \n{json.dumps(json.loads(response.raw.content), indent=4, ensure_ascii=False)}")
        return

    lark.logger.info(lark.JSON.marshal(response.data, indent=4))

def 更新飞书表格(应用ID, 应用密匙, DWBG_TOKEN, DWBG_TABLE_ID, 行ID, 上传数据结构):
    """更新飞书多维表格指定行数据（增强错误日志）"""
    client = lark.Client.builder() \
        .app_id(应用ID) \
        .app_secret(应用密匙) \
        .log_level(lark.LogLevel.DEBUG) \
        .build()

    request: UpdateAppTableRecordRequest = UpdateAppTableRecordRequest.builder() \
        .app_token(DWBG_TOKEN) \
        .table_id(DWBG_TABLE_ID) \
        .record_id(行ID) \
        .request_body(AppTableRecord.builder()
                      .fields(上传数据结构)
                      .build()) \
        .build()

    response: UpdateAppTableRecordResponse = client.bitable.v1.app_table_record.update(request)

    if not response.success():
        error_detail = f"更新表格失败 - 行ID:{行ID} | 代码:{response.code} | 消息:{response.msg} | 日志ID:{response.get_log_id()}"
        if response.raw and response.raw.content:
            try:
                resp_json = json.loads(response.raw.content)
                error_detail += f"\n详细响应: {json.dumps(resp_json, indent=2, ensure_ascii=False)}"
            except:
                error_detail += f"\n响应内容: {response.raw.content}"
        lark.logger.error(error_detail)
        raise Exception(error_detail)  # 抛出异常，让脚本终止并提示错误
    else:
        print(f"✅ 行ID [{行ID}] 更新成功，更新字段: {list(上传数据结构.keys())}")

def 获取多维表格内容(tenant_access_token, app_token, table_id):
    """获取多维表格所有记录（增强错误处理）"""
    all_records = []
    page_token = ''
    has_more = True

    while has_more:
        url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records/search"
        payload = json.dumps({
            "page_token": page_token,
            "page_size": 100
        })
        headers = {
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {tenant_access_token}'
        }

        try:
            response = requests.post(url, headers=headers, data=payload, timeout=10)
            response.raise_for_status()
            result = response.json()

            if result.get('code') != 0:
                error_details = {
                    "code": result.get('code'),
                    "msg": result.get('msg'),
                    "app_token": app_token,
                    "table_id": table_id,
                    "url": url
                }
                raise Exception(f"飞书API错误: {json.dumps(error_details, ensure_ascii=False)}")

            data = result.get('data', {})
            items = data.get('items', [])
            all_records.extend(items)
            has_more = data.get('has_more', False)
            page_token = data.get('page_token', '')

        except requests.exceptions.HTTPError as e:
            raise Exception(f"HTTP请求错误: {str(e)}，URL: {url}，可能是app_token或table_id错误")
        except Exception as e:
            raise Exception(f"获取表格内容失败: {str(e)}")

    return all_records

def 获取多维表格中附件的链接(访问令牌, DWBG_TOKEN, DWBG_TABLE_ID, 行ID=None, 附件字段名="上传附件"):
    """提取多维表格指定行的Excel附件原始URL"""
    if not 行ID:
        raise ValueError("❌ 行ID不能为空，请传入目标行的record_id")

    url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{DWBG_TOKEN}/tables/{DWBG_TABLE_ID}/records/search"
    headers = {
        "Authorization": f"Bearer {访问令牌}",
        "Content-Type": "application/json"
    }
    request_data = {"page_size": 100, "page_token": ""}
    all_attachments = []

    while True:
        try:
            resp = requests.post(url, headers=headers, json=request_data, timeout=15)
            resp.raise_for_status()
            result = resp.json()
        except requests.exceptions.RequestException as e:
            raise Exception(f"❌ 请求接口失败: {str(e)}")

        if result["code"] != 0:
            raise Exception(f"❌ 读取表格失败: {result['msg']} (code: {result['code']})")

        # 精准匹配目标行ID
        target_record = None
        for record in result["data"]["items"]:
            if record["record_id"] == str(行ID):
                target_record = record
                break

        if target_record:
            fields = target_record.get("fields", {})
            attachments = fields.get(附件字段名, [])
            if not attachments:
                raise Exception(f"❌ 行ID [{行ID}] 的「{附件字段名}」列无附件")

            # 筛选Excel格式附件
            for att in attachments:
                att_url = att.get("url")
                att_name = att.get("name", "")
                if att_url and att_name.endswith((".xlsx", ".xls")):
                    print(f"✅ 行ID [{行ID}] 提取到Excel附件: {att_name} | URL前50位: {att_url[:50]}...")
                    all_attachments.append((att_url, att_name))
            break

        if not result["data"].get("has_more"):
            break
        request_data["page_token"] = result["data"]["page_token"]

    if not all_attachments:
        raise Exception(f"❌ 行ID [{行ID}] 的「{附件字段名}」列未找到Excel附件")

    return all_attachments

def 在线解析表格文件(访问令牌, 文件临时链接, 文件名称):
    """在线解析多Sheet的Excel文件（过滤指定Sheet + 跳过空Sheet）"""
    headers = {
        "Authorization": f"Bearer {访问令牌}",
        "User-Agent": "Mozilla/5.0 (Linux; x86_64) AppleWebKit/537.36"
    }

    # 分块读取Excel二进制流（适配大文件）
    resp = requests.get(文件临时链接, headers=headers, timeout=300, stream=True)
    resp.raise_for_status()
    excel_content = b""
    for chunk in resp.iter_content(chunk_size=1024*1024):
        excel_content += chunk

    try:
        excel_file = io.BytesIO(excel_content)
        # 获取所有Sheet名称
        if 文件名称.endswith(".xls"):
            import xlrd
            workbook = xlrd.open_workbook(file_contents=excel_content)
            sheet_names = workbook.sheet_names()
        else:
            from openpyxl import load_workbook
            workbook = load_workbook(excel_file, read_only=True)
            sheet_names = workbook.sheetnames

        print(f"✅ 检测到Excel包含 {len(sheet_names)} 个Sheet: {sheet_names}")

        # 过滤指定关键词的Sheet
        FILTER_KEYWORDS = ["监测数据", "正式数据", "IQA检测"]
        filtered_sheets = [
            name for name in sheet_names
            if any(kw in name for kw in FILTER_KEYWORDS)
        ]
        print(f"✅ 过滤后需解析的Sheet: {filtered_sheets}")

        if not filtered_sheets:
            raise Exception(f"❌ 无符合过滤条件的Sheet（关键词：{FILTER_KEYWORDS}）")

        # 遍历过滤后的Sheet，跳过空Sheet
        all_sheets_data = []
        for sheet_name in filtered_sheets:
            if 文件名称.endswith(".xls"):
                df_sheet = pd.read_excel(
                    excel_file,
                    engine="xlrd",
                    sheet_name=sheet_name
                )
            else:
                excel_file.seek(0)
                df_sheet = pd.read_excel(
                    excel_file,
                    engine="openpyxl",
                    sheet_name=sheet_name
                )

            df_sheet = df_sheet.fillna("")

            if len(df_sheet) == 0:
                print(f"⚠️ 跳过空Sheet: {sheet_name}（0行数据）")
                continue

            df_sheet["所属Sheet"] = sheet_name
            print(f"  - 解析Sheet [{sheet_name}]: {len(df_sheet)}行 × {len(df_sheet.columns)}列")
            all_sheets_data.append(df_sheet)

        if not all_sheets_data:
            raise Exception("❌ 过滤后所有Sheet均为空，无数据可解析")

        df_merged = pd.concat(all_sheets_data, ignore_index=True)
        print(f"\n✅ 所有有效Sheet合并完成: 总计 {len(df_merged)} 行数据")

        return df_merged

    except ImportError as e:
        raise Exception(f"缺少Excel解析依赖: {str(e)}，请安装 xlrd/openpyxl")
    except Exception as e:
        raise Exception(f"多Sheet解析失败: {str(e)}")

def 在线解析表格为二维数据(访问令牌, 文件临时链接, 文件名称):
    """在线解析Excel为{工作表名: 二维列表}字典（无需落地文件）"""
    headers = {
        "Authorization": f"Bearer {访问令牌}",
        "User-Agent": "Mozilla/5.0 (Linux; x86_64) AppleWebKit/537.36"
    }
    try:
        resp = requests.get(文件临时链接, headers=headers, timeout=300, stream=True)
        resp.raise_for_status()
        excel_content = io.BytesIO()
        for chunk in resp.iter_content(chunk_size=1024*1024):
            if chunk:
                excel_content.write(chunk)
        excel_content.seek(0)
        print(f"✅ 成功下载在线附件: {文件名称}")
    except Exception as e:
        print(f"❌ 下载在线附件失败: {str(e)}")
        return None

    # 优先用pyexcel解析
    try:
        工作表字典 = {}
        book = pyexcel.get_book(
            file_type=文件名称.split('.')[-1],
            file_content=excel_content.getvalue()
        )

        for sheet_name in book.sheet_names():
            二维列表 = book[sheet_name].rows()
            二维列表 = [[cell if cell is not None else "" for cell in row] for row in 二维列表]
            工作表字典[sheet_name] = 二维列表

        print(f"✅ pyexcel解析完成，共{len(工作表字典)}个Sheet")
        return 工作表字典
    except Exception as e:
        print(f"⚠️ pyexcel解析失败，降级用pandas: {str(e)}")
        excel_content.seek(0)

    # pandas降级解析
    try:
        工作表字典 = {}
        excel_file = pd.ExcelFile(excel_content)

        for sheet_name in excel_file.sheet_names:
            engine = "xlrd" if 文件名称.lower().endswith('.xls') else "openpyxl"
            df = pd.read_excel(excel_file, sheet_name=sheet_name, engine=engine)

            header_row = df.columns.tolist()
            data_rows = df.values.tolist()
            二维列表 = [header_row] + data_rows

            # 统一处理空值和numpy类型
            二维列表 = [
                [
                    "" if pd.isna(cell)
                    else str(cell) if not isinstance(cell, (str, int, float, bool))
                    else cell
                    for cell in row
                ]
                for row in 二维列表
            ]
            工作表字典[sheet_name] = 二维列表

        print(f"✅ pandas解析完成，共{len(工作表字典)}个Sheet")
        return 工作表字典
    except Exception as e:
        print(f"❌ pandas解析失败: {str(e)}")
        return None

def 日期单元格转变(批次):
    """统一日期格式转换"""
    if isinstance(批次, datetime):
        return 批次.strftime("%Y-%m-%d")
    elif isinstance(批次, str):
        try:
            return 批次[:4] + "-" + 批次[5:7] + "-" + 批次[8:10]
        except:
            return "格式错误"
    else:
        try:
            return datetime.fromtimestamp(批次).strftime("%Y%m%d")
        except:
            return "不支持的格式"

def 判断品项(内容):
    """根据内容匹配产品品项"""
    内容_str = str(内容)
    if "速冻调理九块鸡" in 内容_str:
        return "速冻调理九块鸡2.0"
    elif "冷冻调味鸡架" in 内容_str:
        return "冷冻调味鸡架"
    elif "速冻调理烤翅用鸡翅中尖" in 内容_str:
        return "速冻调理烤翅用鸡翅中尖2.0"
    elif "速冻调理烤翅用鸡翅根2.0" in 内容_str:
        return "速冻调理烤翅用鸡翅根2.0"
    elif "速冻调理辣翅用鸡翅根2.0" in 内容_str:
        return "速冻调理辣翅用鸡翅根2.0"
    elif "速冻调理鸡翅边肉" in 内容_str:
        return "速冻调理鸡翅边肉2.0"
    elif "速冻调理115汉堡用鸡腿肉2.0" in 内容_str:
        return "速冻调理115汉堡用鸡腿肉2.0"
    elif "速冻调理90汉堡用鸡腿肉（香辣）2.0" in 内容_str:
        return "速冻调理90汉堡用鸡腿肉（香辣）2.0"
    elif "速冻调理90汉堡用鸡腿肉（ETC）" in 内容_str:
        return "速冻调理90汉堡用鸡腿肉（ETC）2.0"
    elif "冷冻烧烤风味带皮鸡腿肉" in 内容_str:
        return "冷冻烧烤风味带皮鸡腿肉"
    elif "冷冻原味鸡风味带皮鸡腿肉" in 内容_str:
        return "冷冻原味鸡风味带皮鸡腿肉"
    elif "速冻调理烤翅用翅中尖" in 内容_str:
        return "速冻调理烤翅用鸡翅中尖"
    elif "速冻调理鸡腿肉丁" in 内容_str:
        return "速冻调理鸡腿肉丁2.0"
    elif "速冻调理鸡腿肉条" in 内容_str:
        return "速冻调理鸡腿肉条2.0"
    elif "速冻调理辣翅用鸡翅中2.0" in 内容_str:
        return "速冻调理辣翅用鸡翅中2.0"
    elif "冷冻腌制香辣风味鸡腿肉" in 内容_str:
        return "冷冻腌制香辣风味鸡腿肉"
    elif "火塘烧烤风味翅中尖" in 内容_str:
        return "火塘烧烤风味翅中尖"
    elif "九块鸡" in 内容_str:
        return "九块鸡"
    elif "115汉堡腿肉" in 内容_str:
        return "115汉堡腿肉"
    elif "90汉堡腿肉" in 内容_str:
        return "90汉堡腿肉"
    elif "腿肉条" in 内容_str:
        return "腿肉条"
    elif "腿肉丁" in 内容_str:
        return "腿肉丁"
    elif "翅边肉" in 内容_str:
        return "鸡翅边肉"
    elif "烤翅用翅中尖" in 内容_str:
        return "烤翅用翅中尖"
    elif "烤翅用翅根" in 内容_str:
        return "烤翅用翅根"
    elif "辣翅用翅根" in 内容_str:
        return "辣翅用翅根"
    elif "辣翅用翅中" in 内容_str:
        return "辣翅用翅中"
    elif "鸡翅尖" in 内容_str:
        return "鸡翅尖"
    elif "琵琶腿" in 内容_str and "140" in 内容_str:
        return "琵琶腿（140）"
    elif "琵琶腿" in 内容_str and "110" in 内容_str:
        return "琵琶腿（110）"
    elif "大翅根" in 内容_str:
        return "大翅根（60-70）"
    else:
        return None

# ====================== 工厂名称映射字典 ======================
检查工厂字典 = {
    '光泽二厂': ['福建圣农发展股份有限公司中坊第二肉鸡加工厂'],
    '光泽三厂': ['福建圣农发展股份有限公司中坊第三肉鸡加工厂'],
    '光泽四厂': ['福建圣农发展股份有限公司中坊第四肉鸡加工厂'],
    '光泽六厂': ['福建圣农发展股份有限公司肉鸡加工六厂'],
    '浦城一厂': [
        '福建圣农发展（浦城）有限公司一厂',
        '福建圣农发展（浦城）有限公司肉鸡加工一厂'
    ],
    '浦城二厂': [
        '福建圣农发展（浦城）有限公司肉鸡加工二厂',
        '福建圣农发展（浦城）有限公司二厂'
    ],
    '政和工厂': ['圣农发展（政和）有限公司'],
    '圣越工厂': ['甘肃圣越农牧发展有限公司']
}
新检查工厂字典 = {值: 键 for 键, 值列表 in 检查工厂字典.items() for 值 in 值列表}

# ====================== 核心业务逻辑 ======================
def main():
    """主执行函数（适配GitHub Actions）"""
    # 1. 校验环境变量
    print("🔍 开始校验环境变量...")
    validate_environment()
    print("✅ 环境变量校验通过")

    # 2. 获取飞书访问令牌
    print("\n🔍 开始获取飞书访问令牌...")
    try:
        访问令牌 = 获取访问令牌(APP_ID, APP_SECRET)
    except Exception as e:
        raise Exception(f"获取访问令牌失败: {str(e)}")

    # 3. 获取多维表格附件链接
    print(f"\n🔍 开始提取行ID [{ROW_ID}] 的附件链接...")
    try:
        获取信息 = 获取多维表格中附件的链接(访问令牌, DWBG_TOKEN, DWBG_TABLE_ID, ROW_ID, "上传附件")
    except Exception as e:
        raise Exception(f"获取附件链接失败: {str(e)}")

    # 4. 解析Excel并构建数据字典
    print("\n🔍 开始解析Excel附件并提取数据...")
    数据字典 = {}
    for 列表元素_元组 in 获取信息:
        文件临时链接 = 列表元素_元组[0]
        文件名称 = 列表元素_元组[1]
        print(f"\n📄 处理附件: {文件名称}")
        print(f"🔗 附件链接: {文件临时链接[:50]}...")

        # 解析Excel为二维数据
        读取数据字典 = 在线解析表格为二维数据(访问令牌, 文件临时链接, 文件名称)
        if not 读取数据字典:
            raise Exception(f"附件 {文件名称} 解析失败，返回空数据")

        # 处理"监测数据"工作表
        if "监测数据" in 读取数据字典:
            print(f"📊 开始处理「监测数据」工作表...")
            工作表内容 = 读取数据字典["监测数据"]
            取值字典 = 数据字典.setdefault("监测数据", {})

            for 行数, 一行内容 in enumerate(工作表内容):
                # 跳过表头和空行（行数>1 且 第14列（索引13）有值）
                if 行数 > 1 and 一行内容[13]:
                    # 匹配工厂名称
                    工厂名称 = 新检查工厂字典.get(一行内容[1])
                    # 匹配产品品项
                    产品品项 = 判断品项(一行内容[3])

                    # 跳过无法匹配的工厂/品项
                    if not 产品品项 or not 工厂名称:
                        print(f"⚠️ 行数{行数} - 工厂/品项匹配失败: 工厂={一行内容[1]}, 品项={一行内容[3]}，跳过")
                        continue

                    # 提取核心字段
                    工艺品类 = 一行内容[2]
                    模块 = 一行内容[4]
                    工序 = 一行内容[5]
                    控制组 = 一行内容[6]
                    控制点 = 一行内容[7]
                    检测时间 = 日期单元格转变(一行内容[9])
                    状态 = 一行内容[13]
                    检测值 = 一行内容[14]
                    控制标准 = 一行内容[15]

                    # 筛选条件：生产过程监测 → 原料鸡肉 → 产品品质检查
                    if 模块 == "生产过程监测" and 工艺品类 == "原料鸡肉" and 工序 == "产品品质检查":
                        # 构建嵌套数据字典
                        取值列表 = 取值字典.setdefault(工厂名称, {}) \
                                            .setdefault(检测时间[:10], {}) \
                                            .setdefault(产品品项, {}) \
                                            .setdefault(状态, {}) \
                                            .setdefault(str(控制组), {}) \
                                            .setdefault(str(控制点), [])
                        取值列表.append(检测值)

    # 5. 构建偏差统计汇总字典和翅类中值统计汇总字典
    print("\n🔍 开始统计偏差和翅类中值数据...")
    偏差统计汇总字典 = {}
    翅类中值统计汇总字典 = {}

    for 项, 嵌入字典 in 数据字典.items():
        for 工厂名称, 嵌入字典2 in 嵌入字典.items():
            for 监测时间, 嵌入字典3 in 嵌入字典2.items():
                for 产品品项, 嵌入字典4 in 嵌入字典3.items():
                    偏差明细列表 = []
                    # 统计不合格偏差
                    for 状态, 嵌入字典5 in 嵌入字典4.items():
                        if 状态 == "不合格":
                            for 控制组, 嵌入字典6 in 嵌入字典5.items():
                                for 控制点, 录入值 in 嵌入字典6.items():
                                    偏差明细列表.append(str(控制点))
                    
                    # 构建偏差信息
                    偏差信息 = "、".join(偏差明细列表) if 偏差明细列表 else ""
                    偏差统计汇总字典.setdefault(工厂名称, {}).setdefault(监测时间, {}).setdefault(产品品项, 偏差信息)

                    # 统计翅类单枚重量
                    if "翅中" in 产品品项 or "翅根" in 产品品项:
                        for 状态, 嵌入字典5 in 嵌入字典4.items():
                            for 控制组, 嵌入字典6 in 嵌入字典5.items():
                                for 控制点, 录入值 in 嵌入字典6.items():
                                    if "单枚重量" in 控制点:
                                        监测值 = "、".join(录入值).replace(",", "、")
                                        翅类中值统计汇总字典.setdefault(工厂名称, {}).setdefault(监测时间, {}).setdefault(产品品项, 监测值)

    # 6. 更新偏差数据到飞书表格
    print("\n🔍 开始更新偏差数据到飞书表格...")
    for 工厂名称, 嵌入字典2 in 偏差统计汇总字典.items():
        上传数据结构2 = {}
        一个工厂的数据 = []
        for 监测时间, 嵌入字典3 in 嵌入字典2.items():
            for 产品品项, 偏差信息 in 嵌入字典3.items():
                一个工厂的数据.append(f"{监测时间}*{工厂名称}*{产品品项}*{偏差信息}")
        
        if 一个工厂的数据:
            偏差合并信息 = ",".join(一个工厂的数据)
            字段名 = f"{工厂名称}（偏差）"
            上传数据结构2[字段名] = 偏差合并信息
        
        if 上传数据结构2:
            print(f"📤 准备更新[{工厂名称}]偏差数据: {字段名} = {偏差合并信息[:50]}...")
            更新飞书表格(APP_ID, APP_SECRET, DWBG_TOKEN, DWBG_TABLE_ID, ROW_ID, 上传数据结构2)

    # 7. 更新翅类中值数据到飞书表格
    print("\n🔍 开始更新翅类中值数据到飞书表格...")
    for 工厂名称, 嵌入字典2 in 翅类中值统计汇总字典.items():
        上传数据结构2 = {}
        一个工厂的数据 = []
        for 监测时间, 嵌入字典3 in 嵌入字典2.items():
            for 产品品项, 录入值 in 嵌入字典3.items():
                一个工厂的数据.append(f"{监测时间}*{工厂名称}*{产品品项}*{录入值}")
        
        if 一个工厂的数据:
            翅类中值合并信息 = ",".join(一个工厂的数据)
            字段名 = f"{工厂名称}（翅类中值）"
            上传数据结构2[字段名] = 翅类中值合并信息
        
        if 上传数据结构2:
            print(f"📤 准备更新[{工厂名称}]翅类中值数据: {字段名} = {翅类中值合并信息[:50]}...")
            更新飞书表格(APP_ID, APP_SECRET, DWBG_TOKEN, DWBG_TABLE_ID, ROW_ID, 上传数据结构2)

    # 8. 执行完成
    print("\n🎉 所有数据处理完成！")

# ====================== 脚本入口 ======================
if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ 脚本执行失败: {str(e)}")
        exit(1)  # 退出码非0，标记GitHub Actions任务失败
